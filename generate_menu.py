from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Menu"

# Colors
HEADER_BG    = "1B3A2D"  # dark green
HEADER_FG    = "FFFFFF"
CAT_IDLI_BG  = "FFF3CD"  # warm yellow
CAT_DOSA_BG  = "D4EDDA"  # soft green
CAT_BIRY_BG  = "D1ECF1"  # light blue
CAT_FG       = "2C3E50"
ROW_ODD      = "FAFAFA"
ROW_EVEN     = "F0F0F0"
PRICE_FG     = "1B5E20"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header
headers = ["Category", "Item Name", "Price"]
header_fill = PatternFill("solid", fgColor=HEADER_BG)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color=HEADER_FG, size=12)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

data = [
    ("Sunrise Bites - Mini Idli", "Kumbakonam Ghee Mini Idli",       13, CAT_IDLI_BG),
    ("Sunrise Bites - Mini Idli", "Tanjore Sambar Mini Idli",         13, CAT_IDLI_BG),
    ("Sunrise Bites - Mini Idli", "Chettinad Pepper Mini Idli",       14, CAT_IDLI_BG),
    ("Sunrise Bites - Mini Idli", "Kongu Podi Mini Idli",             13, CAT_IDLI_BG),
    ("Sunrise Bites - Mini Idli", "Tirunelveli Ghee Podi Mini Idli",  14, CAT_IDLI_BG),
    ("Sunrise Crisp - Dosa",      "Palakkad Set Dosa",                14, CAT_DOSA_BG),
    ("Sunrise Crisp - Dosa",      "Salem Podi Dosa",                  14, CAT_DOSA_BG),
    ("Sunrise Crisp - Dosa",      "Pondicherry Cheese Dosa",          15, CAT_DOSA_BG),
    ("Sunrise Crisp - Dosa",      "Turmeric Carrot Dosa",             14, CAT_DOSA_BG),
    ("Sunrise Crisp - Dosa",      "Onion Dosa",                       14, CAT_DOSA_BG),
    ("Sunrise Crisp - Dosa",      "Coimbatore Onion Pepper Dosa",     15, CAT_DOSA_BG),
    ("Garden Biryani - After 2 PM", "Palani Jackfruit Biryani",       16, CAT_BIRY_BG),
    ("Garden Biryani - After 2 PM", "Dindigul Seeraga Samba Biryani", 14, CAT_BIRY_BG),
    ("Garden Biryani - After 2 PM", "Ambur Soya Biryani (Seeraga Samba)", 15, CAT_BIRY_BG),
]

for i, (cat, item, price, bg) in enumerate(data, 2):
    fill = PatternFill("solid", fgColor=bg)

    c1 = ws.cell(row=i, column=1, value=cat)
    c1.fill = fill
    c1.font = Font(bold=True, color=CAT_FG, size=11)
    c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c1.border = border

    c2 = ws.cell(row=i, column=2, value=item)
    c2.fill = fill
    c2.font = Font(color=CAT_FG, size=11)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = border

    c3 = ws.cell(row=i, column=3, value=price)
    c3.fill = fill
    c3.font = Font(bold=True, color=PRICE_FG, size=11)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.number_format = '"$"#,##0'
    c3.border = border

# Column widths & row heights
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 10
ws.row_dimensions[1].height = 28
for i in range(2, 16):
    ws.row_dimensions[i].height = 22

# Freeze header row
ws.freeze_panes = "A2"

wb.save("/home/user/bmad-marketplace/menu.xlsx")
print("Done")
